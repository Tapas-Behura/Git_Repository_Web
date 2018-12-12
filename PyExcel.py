import openpyxl 

# Give the location of the file 
path = "C:\\Python\\Example\\CopyFileExcel.xlsx"

# workbook object is created 
wb= openpyxl.load_workbook(path) 
# WS = wb['SampleSheet']
WS=wb['Sheet1']
Datalist = list(WS.values)

M_Row = WS.max_row
M_Col = WS.max_column


DataHeader = list(Datalist[0])

DataHeader.append('Profit')

print(DataHeader)
